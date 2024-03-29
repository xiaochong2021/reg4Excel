from PyQt5.QtCore import QUrl, QThread, pyqtSignal,QCoreApplication,Qt
from PyQt5.QtWebChannel import QWebChannel
from PyQt5.QtWebEngineWidgets import QWebEngineView
from PyQt5.QtWidgets import QFileDialog, QApplication
from openpyxl import load_workbook, Workbook
import re
import time
import math

from MyObjectCls import MyObjectCls


class MainWin(QWebEngineView):
    reg_path = ''
    text_path = ''
    reg_wb = None
    reg_sh = None
    reg_columns = list()
    text_wb = None
    text_sh = None
    text_columns = list()
    thread = None

    def __init__(self, main_entry):
        QWebEngineView.__init__(self)
        self.setWindowTitle("Excel批量正则匹配工具")
        self.__channel = QWebChannel(self.page())
        self.__my_object = MyObjectCls(self)
        self.__my_object.sigGetRegFilePath.connect(self.getRegFilePath)
        self.__my_object.sigInitPage2.connect(self.init_page_2)
        self.__my_object.sigStart.connect(self.start)
        self.__channel.registerObject('MyObject', self.__my_object)
        self.page().setWebChannel(self.__channel)
        html_url = QUrl.fromLocalFile(main_entry)
        print(html_url)
        self.page().load(html_url)

    def getRegFilePath(self, file_type):
        try:
            f_name = QFileDialog.getOpenFileName(self, '请选择原数据excel文件', os.getcwd(),
                                                 'Excel文件(*.xlsx;*.xls;*.csv;*.xlsm)')
            if file_type == 'reg':
                self.__my_object.sigSetVueRegFilePath.emit(f_name[0], 'reg')
                self.reg_path = f_name[0]
            elif file_type == 'text':
                self.__my_object.sigSetVueRegFilePath.emit(f_name[0], 'text')
                self.text_path = f_name[0]
        except Exception as e:
            print(e)

    def init_page_2(self):
        # 获取正则列
        self.reg_wb = load_workbook(self.reg_path)
        self.reg_sh = self.reg_wb.worksheets[0]
        self.reg_columns = []
        for cell in self.reg_sh['1']:
            self.reg_columns.append({'value': cell.column, 'label': cell.value})

        # 获取文本列
        self.text_wb = load_workbook(self.text_path)
        self.text_sh = self.text_wb.worksheets[0]

        self.text_columns = []
        for cell in self.text_sh['1']:
            self.text_columns.append({'value': cell.column, 'label': cell.value})
        self.__my_object.sigSetVueRegColumns.emit(self.text_columns, self.reg_columns)

    def start(self, text_column_index, reg_column_index, logic_code):
        self.thread = ProcessReg()
        self.thread.set_params(self.reg_sh, self.text_sh, text_column_index, reg_column_index, logic_code)
        self.thread.trigger.connect(self.message_ui)
        self.thread.start()

    def message_ui(self, message, message_type):
        if message_type == 'spin':
            self.__my_object.sigLoadingTip.emit(message)
        else:
            self.__my_object.sigInfo.emit(message, message_type)


class ProcessReg(QThread):
    trigger = pyqtSignal(str, str)
    reg_sh = None
    text_sh = None
    text_column_index = None
    reg_column_index = None
    logic_code = None

    def __init__(self):
        super(ProcessReg, self).__init__()

    def set_params(self, reg_sh, text_sh, text_column_index, reg_column_index, logic_code):
        self.reg_sh = reg_sh
        self.text_sh = text_sh
        self.text_column_index = text_column_index
        self.reg_column_index = reg_column_index
        self.logic_code = logic_code

    def check_text(self, search_text, logic_code, reg_column_index):
        # 匹配到的正则组名称
        reg_match_list = list()
        # 逻辑码表中所包含的数字
        logic_num_list = re.findall(r'\d+', logic_code)

        for row_index in range(1, self.reg_sh.max_row + 1):
            if row_index == 1:
                continue
            # 转换码表逻辑
            re_logic_map = dict()
            for logic_num in logic_num_list:
                if logic_num not in re_logic_map:
                    # excel的数据有可能时数字，需要转成字符串
                    reg_cell_value = str(self.reg_sh.cell(row=row_index, column=int(logic_num)).value)
                    math_str = re.search(r"" + reg_cell_value, search_text, re.I)
                    re_logic_map[logic_num] = str(math_str is not None)

            logic_code_list = re.split(r'\s+|\b(?=\()|\b(?=\))|(?<=\()\b|(?<=\))\b', logic_code)
            for index, code in enumerate(logic_code_list):
                if code in re_logic_map:
                    logic_code_list[index] = re_logic_map[code]
            logic_code_ts = ' '.join(logic_code_list)

            try:
                result = eval(logic_code_ts)
                if result:
                    reg_match_list.append(self.reg_sh.cell(row=row_index, column=reg_column_index).value)
            except Exception as e:
                return False

        return '，'.join(reg_match_list)

    def run(self):
        try:
            result_wb = Workbook()
            result_sh = result_wb.active
            header_row = list()
            for cell in self.text_sh[1]:
                header_row.append(cell.value)
            header_row.append(self.reg_sh.cell(row=1, column=self.reg_column_index).value)
            result_sh.append(header_row)
            is_logic_code_error = self.check_text('测试', self.logic_code, self.reg_column_index)
            if type(is_logic_code_error) is bool:
                self.trigger.emit('码表逻辑语法错误', 'error')
                return
            progress_num = 0
            text_columns_tuple = tuple(self.text_sh.columns)
            for cell in text_columns_tuple[self.text_column_index - 1]:
                if cell.row == 1:
                    continue
                result = self.check_text(str(cell.value), self.logic_code, self.reg_column_index)

                result_row = list()
                for cell_item in self.text_sh[cell.row]:
                    result_row.append(cell_item.value)
                result_row.append(result)

                result_sh.append(result_row)
                progress_num_now = math.ceil(cell.row / self.text_sh.max_row * 100)
                if progress_num != progress_num_now:
                    self.trigger.emit('{}%'.format(progress_num_now), 'spin')
                    progress_num = progress_num_now
            result_wb.save(time.strftime("导出文件/%Y-%m-%d-%H%M%S.xlsx", time.localtime()))
            self.trigger.emit("匹配完成，结果已导出", "success")
        except Exception as e:
            self.trigger.emit("未知错误", "error")


if __name__ == '__main__':
    import sys, os

    QCoreApplication.setAttribute(Qt.AA_EnableHighDpiScaling)
    app = QApplication(sys.argv)
    main_entry = os.path.realpath(os.path.dirname(__file__) + "/content/index.html")
    print(main_entry)
    w = MainWin(main_entry)
    w.resize(600, 500)
    w.show()
    sys.exit(app.exec_())
